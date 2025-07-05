from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from tkinter import messagebox
from datetime import datetime
import pandas as pd
import base64
import tempfile
import os, io, base64

class SAP:

    #constructor - initializing the values
    def __init__(self):
        
        self.user_name = ""
        self.input_cols = [
                                '# of SAP Basis Production Systems in the Landscape with 9x5',
                                '# of SAP Basis Non-Prod Systems in the Landscape with 9x5',
                                '# of SAP Basis Production Systems in the Landscape with 16x5',
                                '# of SAP Basis Non-Prod Systems in the Landscape with 16x5',
                                '# of SAP Basis Production Systems in the Landscape with 24x7',
                                '# of SAP Basis Non-Prod Systems in the Landscape with 24x7',
                                '# of SAP Transport Requests - Production Systems per month',
                                '# of SAP Transport Requests - Non - Production Systems per month',
                                '# of SAP Basis P1/P2 Incidents per month',
                                '# of SAP Basis P3/P4/P5 Incidents per month',
                                '# of SAP Basis Changes per month',
                                '# of SAP Basis SRQs per month',
                                '# of SAP Basis Problems per month',
                                '# of SAP maintainance activties per month',
                                '# of SAP System Refreshs Per year',
                                '# of SAP Support Packages per year',
                                '# of SAP Kernel Upgrades per year',
                                '# of SAP Basis Support in Internal / External audits per year',
                                'Misc(Conference Calls, Documentation)',

                                '# of SAP OPM in FTE',
                                '# of SAP SC-OPM in FTE',
                                '# of SAP CDM/SDM in FTE',
                                '# of SAP CSM in FTE'
                         ]
        

        self.volume = [0] * 19 + ['FTE'] * 23
        self.complexity_matrix = [
                                    [0.2, 0.25, 0.3],
                                    [0.1, 0.15, 0.2],
                                    [0.25, 0.28, 0.35],
                                    [0.12, 0.17, 0.25],
                                    [0.28, 0.3, 0.38],
                                    [0.15, 0.2, 0.28],

                                    [0.2, 0.2, 0.2],
                                    [0.15, 0.15, 0.15],
                                    [2, 2.5, 3],
                                    [0.5, 1, 1.5],
                                    [1, 2, 4],
                                    [2, 3, 4],
                                    [8, 12, 16],
                                    [4, 8, 12],
                                    [24, 32, 40],
                                    [8, 12, 16],
                                    [2, 2, 2],
                                    [40, 80, 120],
                                    [8, 24, 40],

                                    [0.1, 0.2, 0.25],
                                    [0.1, 0.15, 0.2],
                                    [0.1, 0.2, 0.25],
                                    [0.05, 0.1, 0.15]
                          ]
        
        self.total_efforts = [0] * 23
        self.mode = 0
        self.eu = [0] * 23
        self.non_eu = [0] * 23

        self.eu_resource = [0] * 23
        self.eu_rate_card = [0] * 23
        self.non_eu_resource = [0] * 23
        self.non_eu_rate_card = [0] * 23
        
        self.eu_fte = [0] * 23
        self.non_eu_fte = [0] * 23
        self.overall_fte = [0] * 23

        self.eu_cost = [0] * 23
        self.non_eu_cost = [0] * 23
        self.overall_cost = [0] * 23

        self.eu_price = [0] * 23
        self.non_eu_price = [0] * 23
        self.overall_price = [0] * 23

        # for the last four rows, assigning different values based on the classes of them
        # Add this to __init__ in SAP class
        self.fte_rate_cards_eu = [
            {"Junior": 21, "Senior": 38, "Specialist": 28, "Expert": 31},  # Row 19
            {"Junior": 25,  "Senior": 42, "Specialist": 30, "Expert": 34},  # Row 20
            {"Junior": 25,  "Senior": 42, "Specialist": 30, "Expert": 34},  # Row 21
            {"Junior": 25,  "Senior": 42, "Specialist": 30, "Expert": 34 }   # Row 22
        ]

        self.fte_rate_cards_noneu = [
            {"Junior": 11, "Senior": 22, "Specialist": 15, "Expert": 19},      # Row 19
            {"Junior": 13, "Senior": 24, "Specialist": 18, "Expert": 20},      # Row 20
            {"Junior": 15, "Senior": 26, "Specialist": 18, "Expert": 23},      # Row 21
            {"Junior": 15, "Senior": 26, "Specialist": 18, "Expert": 23}       # Row 22
        ]


    def user_inputs(self):
        # Note : FROM THE '# of SAP OPM Efforts' YOU SHOULD ENTER FTE VALUES
        print()
        self.user_name = input("Customer Name : ")
        for i in range(19):  
            while True:
                try:
                    user_input = input(f"{self.input_cols[i]} : ").strip()
                    self.volume[i] = int(user_input)
                    break 
                except ValueError:
                    print("Invalid input! Please enter a number.")
        print("\n")

        
        for i in range(23):  
            if i >= 19:
                print('\n Enter either (0% or 100%) \n')
            while True:
                try:
                    print(self.input_cols[i]," : ")
                    eu = int(input("Enter EU Percentage: "))
                    non_eu = int(input("Enter Non-EU Percentage: "))
                    print()
                    if eu + non_eu == 100:
                        self.eu[i] = eu
                        self.non_eu[i] = non_eu
                        break
                    else:
                        print(" Sum of EU and Non-EU percentages must be exactly 100. Please try again.\n")
                except ValueError:
                    print(" Invalid input! Please enter valid integer percentages.\n")
        print("\n")


        
        self.mode = int(input("\n0.Simple[0-50 SID'S]  1.Medium[50-100 SID'S]   2.Complex[100+ SID'S] \n Enter the complexity : "))
        temp = {0 : 'SIMPLE', 1 : 'MEDIUM', 2 : 'COMPLEX'}
        if self.mode in temp.keys():
            print("\nSelected ",temp[self.mode]," complexity\n")
        else:
            print("\nEnter the correct complexity as mentioned :\n ")
            self.mode = int(input("\n0.Simple[0-50 SID'S]  1.Medium[50-100 SID'S]   2.Complex[100+ SID'S] \n Enter the complexity : "))

    
#        if input("\nEdit Complexity Values - Type (Yes or No) : ").strip().lower() == 'yes':


        # Edit Complexities
        for i in range(23):
            print(self.input_cols[i]," : ",self.volume[i]," : ",end = "")
            while True:
                try:
                    self.complexity_matrix[i][self.mode] = float(input())
                    break
                except ValueError:
                    print("\nInvalid input, Enter a numeric value\n ")



        #Resource EU
        print("\nJ - Junior(2) || S - Senior(4) || Sp - Specialist(6) || E - Expert(8) \n")
        print("\nChoose the EU Resource person : \n")
        for i in range(23):
            while True:
                    try:
                        self.eu_resource[i] = input().strip().lower()
                        break
                    except ValueError:
                        print("\nInvalid input, Enter correct value\n ")

        #Resource Non-EU
        print("\nJ - Junior(2) || S - Senior(4) || Sp - Specialist(6) || E - Expert(8) \n")
        print("\nChoose the Non-EU Resource person : \n")
        for i in range(23):
            while True:
                    try:
                        self.non_eu_resource[i] = input().strip().lower()
                        break
                    except ValueError:
                        print("\nInvalid input, Enter correct value\n ")


        #Assigning Rate Card values for EU
        for i in range(23):
            if self.eu_resource[i] == "j":
                self.eu_resource[i] = "Junior"
                self.eu_rate_card[i] = 35
    
            elif self.eu_resource[i] == "s":
                self.eu_resource[i] = "Senior"
                self.eu_rate_card[i] = 35

            elif self.eu_resource[i] == "sp":
                self.eu_resource[i] = "Specialist"
                self.eu_rate_card[i] = 35
            else:
                self.eu_resource[i] = "Expert"
                self.eu_rate_card[i] = 35

        #Assigning Rate Card values for Non-EU
        for i in range(23):
            if self.non_eu_resource[i] == "j":
                self.non_eu_resource[i] = "Junior"
                self.non_eu_rate_card[i] = 11.21
    
            elif self.non_eu_resource[i] == "s":
                self.non_eu_resource[i] = "Senior"
                self.non_eu_rate_card[i] = 11.21

            elif self.non_eu_resource[i] == "sp":
                self.non_eu_resource[i] = "Specialist"
                self.non_eu_rate_card[i] = 11.21
            else:
                self.non_eu_resource[i] = "Expert"
                self.non_eu_rate_card[i] = 11.21



    def calculate_efforts(self):
        for i in range(23):
            if i >= 19:
                self.total_efforts[i] = self.complexity_matrix[i][self.mode]
            else:
                self.total_efforts[i] = self.volume[i] * self.complexity_matrix[i][self.mode]
            
            if i in range(14, 18):
                self.total_efforts[i] = round(self.total_efforts[i] / 12, 4)
        

    def final_calculations(self):
        self.fte_summary = 0
        self.cost_summary = 0
        self.price_summary = 0

        for i in range(23):
            self.eu[i] = self.eu[i] / 100
            self.non_eu[i] = self.non_eu[i] / 100

            if i >= 19:            
                self.eu_fte[i] = (self.total_efforts[i] * self.eu[i])
                self.non_eu_fte[i] = (self.total_efforts[i] * self.non_eu[i])
            else:
                self.eu_fte[i] = (self.total_efforts[i] * self.eu[i]) / 140
                self.non_eu_fte[i] = (self.total_efforts[i] * self.non_eu[i]) / 160

            self.overall_fte[i] = self.eu_fte[i] + self.non_eu_fte[i]

            self.eu_cost[i] = self.eu_fte[i] * self.eu_rate_card[i] * 140
            self.non_eu_cost[i] = self.non_eu_fte[i] * self.non_eu_rate_card[i] * 160
            self.overall_cost[i] = self.eu_cost[i] + self.non_eu_cost[i]

            self.eu_price[i] = self.eu_cost[i] / 0.7
            self.non_eu_price[i] = self.non_eu_cost[i] / 0.7
            self.overall_price[i] = self.eu_price[i] + self.non_eu_price[i]

            self.fte_summary += self.overall_fte[i]
            self.cost_summary += self.overall_cost[i]
            self.price_summary += self.overall_price[i]



    def fill_excel(self, region_mode="both"):
        self.dummy_complexity_matrix = [0] * 23
        for i in range(23):
            self.dummy_complexity_matrix[i] = self.complexity_matrix[i][self.mode]
        
        temp = ""
        if self.mode == 0:
            temp = "G"
        elif self.mode == 1:
            temp = "H"
        else:
            temp = "I"

        output_path = f"{self.user_name}_SAP_BASIS_EFFORT_ESTIMATION_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"

        encoded_excel_string = b''' UEsDBBQABgAIAAAAIQBMyoOglQEAAP4FAAATAAgCW0NvbnRlbnRfVHlwZXNdLnhtbCCiBAIooAACAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAADEVNtKAzEQfRf8hyWv0k31QUS69cELCN5A/YDpZrYbmk1CZtT2752NVURqtVjwZZfdZM5lkjmjk3nnimdMZIOv1H45VAX6Ohjrp5V6fLgYHKmCGLwBFzxWaoGkTsa7O6OHRUQqpNpTpVrmeKw11S12QGWI6GWlCakDls801RHqGUxRHwyHh7oOntHzgHsMNR6dYQNPjovzufx+UzKxXhWnb/t6qkpBjM7WwCJUP3vzhWQQmsbWaEL91Al0STEhGGoRuXNlTFYY0z0yizFSeiVnQkebkS5dlVKZhVFrI+2J9W8Y+pXvXS3rbuU4kjVY3EHiG+jEu547/RLSbBLCrFwPsmlrcovKDqx/172GP28mnV/7WxbS+8vAG+o4+CcdLHcddX7+vRUZ5gfjxAuHtO3jz6A/MbeQ0NyzTNF06wI+Y6/TIaN9jQwGGPQVTNBd+ib8ohsdLbOhrB0QWYmJPKuux/gY1lWXXhjvUogkeZVw876/h0NfPYgChIkt/o5Rwu7PB419mho0K7h1Tu/xKwAAAP//AwBQSwMEFAAGAAgAAAAhADEdic0iAQAA3gIAAAsACAJfcmVscy8ucmVscyCiBAIooAACAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAACsktFKAzEQRd8F/yHkvTvbVUSk276IUFAQqR8wTWa3oUkmJFHbvzetFl2oRdDHZO7cnLmTyWzjrHilmAz7Vo6rWgryirXxfSufF3ejaylSRq/RsqdWbinJ2fT8bPJEFnNpSisTkiguPrVylXO4AUhqRQ5TxYF8qXQcHeZyjD0EVGvsCZq6voL43UNOB55irlsZ5/pCisU2lJf/4g2OMmrMCIojjUIsZDGbMotYYOwpt1KzeizXaa+oCrWE40DND0DOqMiJu1wpdsBdZ9RuzKaGuhlOCspiSqYI9glaXJIdkjwccO93tbnv+BTR+PcRfWDdsnpx5PORLXyCHxRf+WwsvHFcL5nXp1gu/5OFNpm8Jn16YRjCgQgGv3L6DgAA//8DAFBLAwQUAAYACAAAACEAkidU38UDAABJCQAADwAAAHhsL3dvcmtib29rLnhtbKxVbW/iOBD+ftL9Byvf09h5gSQqXZG3u0ptt6JseychITcxxSKJOccpVNX+9xsHQtvldOK6h8DG9uTxPDPPTM6/bKsSPTPZcFGPDHKGDcTqXBS8fhoZ36aZ6RuoUbQuaClqNjJeWGN8ufj1l/ONkKtHIVYIAOpmZCyVWoeW1eRLVtHmTKxZDScLISuqYCmfrGYtGS2aJWOqKi0b44FVUV4bO4RQnoIhFgues0TkbcVqtQORrKQK3G+WfN30aFV+ClxF5apdm7mo1gDxyEuuXjpQA1V5ePlUC0kfS6C9JR7aSvgO4EcwDHZ/ExwdXVXxXIpGLNQZQFs7p4/4E2wR8iEE2+MYnIbkWpI9c53Dg1dy8EmvBgeswRsYwT+NRkBanVZCCN4n0byDb7Zxcb7gJbvfSRfR9fqGVjpTpYFK2qi04IoVI2MIS7FhHzZku45aXsKpPQxs37AuDnK+lahgC9qWagpC7uHBENsOxtoShDEuFZM1VSwWtQId7nn9rOY67HgpQOFowv5quWRQWKAv4AojzUP62NxStUStLEdGHM6+NUB/NoZacj3X8fzZ15olkj8zZKKEtUpXI5qykq1Ehca/zRKx6gqHzW5EbUa04Q1KG8WhRkE7sztarUuGYlGwZraP7NyevVM6PS6r/6B1musAWhDBHcvd/x+jCWRl2Ov5VkkE/y+TK8jpHX2GDIOOin0DuIQUEmde5zIk89c48XzbszMzIENiulnqmkESO6bnYDuzk2joBfF3ICMHYS5oq5Z7ihp6ZLiglKOja7rtTwgOW168ufGK9x9Tzz8M/dl3TVi3yXvONs2bzPQSbR94XYjNyDAJhjb78nG56Q4feKGWID/H9qAcd3u/M/60BI+J7epNKCft2ch4tV2MgbBrEnfom64DQxQErplFwZD4Y5ImGek8st651DVkcK2bUd0V0YLXtJx3O9D+dcfuIm0gGeqL5GVBNLH3j9yJstUiemcPLfJgb3eZ7+/KaZlDpempAw4ItgNtwbbqqlHdDCLnwIm4eDzEQAKnjme6fmCbvuvYZuwmduoN0ySNPJ1U/RYK/49e3NVa2L/etJdLKtVU0nwFL8UJW0DZgAq7AFjg73tnI8+PsAMuuhnJTJcE2IyigWt6SeZ4Q5LEqZe9OavpLz7ZCX2re5pR1UKX0A2iW4d6zPa7h83FbmOf3A8FG04SHff90/9meAfsS3aicXZ/omF8cz29PtH2Kp3OH7JTjcfXUTI+3X48mYz/nKZ/9FdY/xjQXcL12MnU6mVy8TcAAAD//wMAUEsDBBQABgAIAAAAIQBKqaZh+gAAAEcDAAAaAAgBeGwvX3JlbHMvd29ya2Jvb2sueG1sLnJlbHMgogQBKKAAAQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAC8ks1qxDAMhO+FvoPRvXGS/lDKOnsphb222wcwsRKHTWxjqT95+5qU7jawpJfQoyQ08zHMZvs59OIdI3XeKSiyHAS62pvOtQpe909X9yCItTO69w4VjEiwrS4vNs/Ya05PZLtAIqk4UmCZw4OUVFscNGU+oEuXxsdBcxpjK4OuD7pFWeb5nYy/NaCaaYqdURB35hrEfgzJ+W9t3zRdjY++fhvQ8RkLyYkLk6COLbKCafxeFlkCBXmeoVyT4cPHA1lEPnEcVySnS7kEU/wzzGIyt2vCkNURzQvHVD46pTNbLyVzsyoMj33q+rErNM0/9nJW/+oLAAD//wMAUEsDBBQABgAIAAAAIQCk1+IvXAwAAM4/AAAYAAAAeGwvd29ya3NoZWV0cy9zaGVldDEueG1snJVfr5owGMbvl+w7kN4LtCCo0XMyNWa7WZadbee6lqrNAepoPeqWffe9LRRMTAjnJGqL8v7eP89TnD9eitx75ZUSslwg7IfI4yWTmSj3C/Tzx2Y0QZ7StMxoLku+QFeu0OPDxw/zs6xe1IFz7QGhVAt00Po4CwLFDrygypdHXsIvO1kVVMNltQ/UseI0s0FFHpAwTIKCihLVhFk1hCF3O8H4WrJTwUtdQyqeUw31q4M4Kkcr2BBcQauX03HEZHEExFbkQl8tFHkFm33Zl7Ki2xz6vuCYMu9SwYvAO3Jp7Pd3mQrBKqnkTvtADuqa79ufBtOAspZ03/8gDI6Dir8KI2CHIu8rCY9bFulg0TthSQsz46pmJ5Et0N8k/USS6SYcLTcYj+J1DB6LV3i0mi4JxsvVejoO/6GHeSZAYdOVV/HdAi3x7DkKUfAwtwb6JfhZ3ew9TbdPPOdMc0iCkfdHyuKJUaPdBNK3l1+NIfP6S+PhrZQvBvYFwkJIqyzEpKVMi1e+4jncvSJwMNRvW4nZQxlBW8ft3tW0sb7/VnkZ39FTrr/L82cu9gcNxcV+DPMwhppl1zVXDJwMyf3IYJnMgQGfXiHgRIIMBb3Y9SwyfbDhk0mS4kk6Rt6WK70RBoo8dlJaFs/NXQ2rpsAALAXWhjJO/DRNozh6AwWqthRYG0rqRxEJI0yGlwJ3WgisDWTik7qZngaSJgrWNgqH07dUnzaIaYvA2GUeOEYMJrDFm42TY+wnSRwmbxgBBrFqDGw6VVtMzxyw84PZuFDi2lD6avwOHfYhnBlw5wYMw+kLccrjTnpM/DjGTePDEjvpcac9jrrp9RXg9MedAaCA8ThO7DHoC3W6mx67iXW+7YuFP79aqBvTkEEVm+dFfXBvrHI7soGWI84r5EZwEK6nauIENpuuY3dc7iMD+8z5DwAA//8AAAD//5SbW2/bRhBG/0qgHxCbkpwbbD/EsmTSDO+XZ8M1mqBoUsRp2v77kt6hODNHQKSnAAffDlc8XJI7Zi6fPz89/dg8/Hi4vvz+7Z9X368W0eLV818PX5+vFssPy9Xi1b/R+uHxw2//bZ6eH5++/rhanL9eLa4vH8fsxzF8tVitF2dCbjQ5G0ru6y5d3c9DqWj9er149fj3849vf949ffn9hf3ikGOdq8V6+Gco8DzM5Of1+eXZz+vLs8dpDiFyca4ikY1sJDLMdl9laSO3EtEHWtnIViKr/c/fBbIcftW+7toOupPIhYpc2Eh8IPLGRhI59BtV5a2N3IfI6p0+kPuRqZxNfare2TKfDpVxJyI7UOa9LZPLhPVsIieukDIX+/NZCnmzJ5UQfYYjZ7dGnQZ1WiFv9UXizk6HOr2uYy7vYaUcv2zG8HANv5uXDcgG5DaQi3mxbYXM52sXiL6sI+fq7teR+NeRRA79dv8T7gNZvd+TVH5CtCefkMmQyYUMZ3q/hCK3hopDGbeIykMZt4qqQxm3jOpDGbdGmkMZtwDaA5mlWwDdoYy7uPtDmfnCNRflsEjMRflyfx1vPJ+/PP7x8dt4Bz58t73Y3+HHEsOlOt/ibkA2ILeBaIVv3anfHoi4M78LkfnQdx7EHiQBXKjLUiY3X5UBROfzVekjGSI5fmIBUoJUIDVIA9KCdCC9Jkb5YPf4+9AYHuS+109Bd7u4CZlIP2iWbjluQmY5n/XbQOYb09aDXQCrOXIHEoMkcqT5JnPvC6cyYXXX8ZEMkRykAClBKpAapAFpQTqQXhPjd/BwvN8x/OJ3/3oWSKSfekt349yEjLYZiLLpwS4AbRMkBknkSMqmL5zKhJVNH8kQyUEKkBKkAqlBGpAWpAPpNTE2Bw/H2xzD1mYg1qa7z25CRtsMRNn0YBeAtgkSgyRyJGXTF05lwsqmj2SI5CAFSAlSgdQgDUgL0oH0mhibw7vt8TbHsLUZiLXpHombkNE2A1E2PdgFoG2CxCCJHEnZ9IVTmbCy6SMZIjlIAVKCVCA1SAPSgnQgvSbG5vBUPN7mGLY2A7HPTffauAmZ1fwqchuIsunBTsao5yZIDJIEslQ2feFUJqxs+kiGSA5SgJQgFUgN0oC0IB1Ir4mxObzvndLXGNP2tfelwPDCPKz+uWXg3vE3EtJGBSmlILtpmJJKFBMlgrRXVE+nmSuzCGUM5UQFUUlUEdVEDVFL1BH1BlnJpzWvQq/KtJJWvpcUhVA0vJDtta98N0lCRnsYp7V7spuGae3SP5tRzFQiyGj31dNp5lq7D2UM5UQFUUlUEdVEDVFL1BH1BlntY1fm+J6l9HDUljYKaHhNmx27XtBGMsax9LT2O8mthFRDZhqmHYdh6oEcM5UIMo798VIJDadmesP/hClkDOVEBVFJVBHVRA1RS9QR9QZZx2Pb43jHB5okK7+zjULIWPc7W8kY62GYXtme7KZh2rr0vvTKBkpkoLHuq6cSMtZ9KGMoJyqISqKKqCZqiFqijqg3yFof+xzHW5euiOqrRgEZx36/KxnjOAzTjj3ZTcO04xAyKxsokYHGsa+eSsg49qGMoZyoICqJKqKaqCFqiTqi3iDreOx1HO84dEZ073x8Ng8va8ax3wVLxjj2baGthPTdO2T03mmqpNcxUomkjGN/vFRCxrEPZQzlRAVRSVQR1UQNUUvUEfUGWcdjB+R4x6FfYhwHZBz7vfG41xr/Gqm2U4L0Og4h7ViG6XUMFE/F51QiyDj21VMJGcc+lDGUExVEJVFFVBM1RC1RR9QbZB2PfZHjHYcuiv7Dwri18uvY75glYxz7FtJWQtpxyJh1DBRPxbXjkDKO/fFSGWcc+1DGUE5UEJVEFVFN1BC1RB1Rb5B1PHZLjncceit2g+W2zDdRCOmVvXabsI1kjHXfatpKSFsPGWMdKJ6Ka+shZaz746Uyzlj3oYyhnKggKokqopqoIWqJOqLeIGt97Kocbz30YMzdOyDjGHvokDGOfQNqG3myE2IcSyX9hAZKZKBx7KunEjKOfShjKCcqiEqiiqgmaohaoo6oN8g4Xp7UH3tJ2y8YBBnHfg8tGe1YkHpCg+ymYeoJTRQTJYK0Y1RPhWjHCGUM5UQFUUlUEdVEDVFL1BH1BlnHJ7XHltIeU7spQcax21VvJGMcoxcmIXWvnoZpx+yFMZUIMo7RC5OQcYxeGEM5UUFUElVENVFD1BJ1RL1B1vFJvbCl9MK044B0m3vtuyIybDW3nW4F6XXsu1W7adj8NdcdUUyUCDKO0QuTkHHsQxlDOVFBVBJVRDVRQ9QSdUS9QdbxSb2w8SPO8WtOs2pdD+RGQibj99CSMStbOlhzL3U7pfRHkehz3U0p9YwmSqapz5fnPcunE1rOTVCmMqZyooKoJKqIaqKGqCXqiHqDrPmT+mHjZ6iD+eGldP4Y1W2abyQztmfmkN91SciYR4dMQvqejt7X3VRJe2eHTFJmvaNDNs1c9b4xhYyhfELz5VIQlRxYEdVEDVFL1BH1BlnrJ3XIlqGFZNc7tMsHRvrPWmv/10ypZLSjaSYhrZ1Ns6mS1s6m2TRz9VdqVE+FmNs8mmYM5dMUtHaZwoxKDqyIaqKGqCXqiHqDrPaTmmajb9zmoT2ExkbdfrVf+N22VDLafctqKyGtnX20qZLWjlQyzVxrRx9NQkY7+mgM5dMUtHaZgtYu52W+l1SsVRM1RC1RR9QbZLWf1EcbPgU6Qrt8smS0479ESENMtU+luH6p832tnWT0BpwoJkqmmWvtaK1JyGhHa42hfDqe1i6/T2uX86K1A9Us3xC1RB1Rb5DVfkprLV+GTtN87gqQEqQCqUEakBakA+k1MT9ruIsc3zu6f0lfLV5+VihzNv8Ppf8BAAD//wAAAP//bJLdboMwDIVfJcoDjOIMdbKAC+gmrS2lFCjXdIQfjRIUUu31B5XYbnwXfz7yOZHt3qVuZCj7fmJf6jEYj4PDffcPMy1rjwe2wMDecovovGEAQHTOAs+vBE8EJhS/CLxQPHBm6w1pbc8dyjoWGJOjAAOKh4AhxXeAO4q/A34IItEecE/pI8CI0p8ATxSPARPqXwfAA6U/Ah4pngvMqTyZwIziqcCUXA1gRuXJAQuKXwVeqTmFwOLJrf+z892xbGRU6qYbJtbLej7BzcuWM9017fo2anxSh7ObMkbd16qVZSX1UgnOaqXMWsyXusxNpXmMTOlODqY0nRo8PiptdNmZ2QG7yuP6s7KXXVo/Sn9PrZTG/wUAAP//AwBQSwMEFAAGAAgAAAAhACiZDrOjBQAAjRgAABgAAAB4bC93b3Jrc2hlZXRzL3NoZWV0Mi54bWyc00mvmzAQAOB7pf4Hy3cWA4GAQp7SoKjvVlWv7dkxJljBmNrOpqr/vYOzVcqh6ZOAMds3HjzMXo6yQ3uujVB9iYkfYsR7pmrRb0r87W3lTTEylvY17VTPS3ziBr/MP36YHZTempZzi0DoTYlba4ciCAxruaTGVwPv4U6jtKQWTvUmMIPmtHYvyS6IwjANJBU9PguFfsZQTSMYrxTbSd7bM6J5Ry3M37RiMFdNsmc4SfV2N3hMyQGIteiEPTkUI8mK102vNF13UPeRJJSho4Ytgj2+pnHXHzJJwbQyqrE+yMF5zo/l50EeUHaTHut/iiFJoPlejAt4p6L3TYlMblZ0x+J3YukNGz+XLnaiLvGvxWqV5cuQeBWJYy/JydL7VOUTj4RptMjScBrmy994PqsFrPBYFdK8KfGCFFUS42A+cw30XfCD+WuMxn5cK7Udb7xCnhAIwzvOxs5AFMKeL3nXlbhKoKV/OrRKiipyaHBT57P7+Jph5br4i0Y1b+ius1/V4TMXm9bCL5P44Lk2KOpTxQ2DvoT0vmOZ6sCAI5Ji/L+grejRxYOobQujyM+yLE7ibILRmhu7EiOKEdsZq+SPy1Nj3TcFlsYpEC/KxI/jKIxJ9DwCS+IQiFck9UmY/89EoHBnQLwYqT+dphmZ/ruawH2aPwAAAP//AAAA//+UmGGO4jAMha+CeoCFJAVmRoAE7UUQi7S/ZlcDYnZvv0nsJvHDpc0PJISf4lfH+eqwu/26Xu/9+X4+7L5+fy++9o1pFrc/58+b//bRNou/pj1fPn7+66+3y/Xzvm9WP1xz2F2C9Bi0XueahQ/c/M+Pw/p9t3wcdsuL//gV07JeM3fZU9D6Zd/LZdu0bEzdkcauSs1aanpNs9Ht+SedbS9o94018XkN+KKg2Ra+LHjvSeO2zRKKtK5wEbTJhQUXFJQusDqkUVxsKlwEbXLhwAUFpYtc/7iPPWkUF76Cs3ckaJML7BQKShdb6BTSKC7eKlwEbXIB1e4oKF28gQvSKC78SZhdi6BNLqDaHQWli3xgaUdIo7gw/rDNthHFyQfUu+Oo8T2YyOHgKPUs0pxUQOpkiFJ8XqHmHUd9G2YjcJp61mhGbE1JgjiVBMreGYoKIwgO1mhGavgagF0QbAXHlsPCCcKDNZqTGpQayVKEKYeFEwQIazQnNTgNrVjUBIHKYeEEIcIazUkNUo1gqkGoclg4QZCwRnNSg9XAiaImCFYOl05a6KWeNZqTGrQawVaDcOWwcPJEk1G6hkFjPtcEXw0CNq61b4STJ5yMEjaMM7OdRHGeQxCxHBZOoJd61ii742lZ4UQQ1iBi41pQEyQbazQnNYi1ArEGGcvh0C55YMVjzCLNSg1krYCsRchyWFrBc8wizUoFUY6WBzEXBs74kj8NP72csVkkhuxNfhAx39sasESxb91wk3gckPpD1O9l2qVNFsm0NRSxTBFKi4gfoiJtFsm0NciwjAxKizwfoiJtFslLVA0fXBCnIj/dnzgq0maRTFtxBI+OjqAfEPB248rj47+/vFpGcbxbpp4dfnp9L1QufZv8vpDPVY4qE4ZOjkcVvWeHqCjnyF3TlXPJZFqeS/SejUv5PRZpM9Dk05bImEzLQ4jes46jIm2Gl0xbgmEyLU8clPapZzkq0o7931CCYTKtAANOF46jZdrtCAZdCYbJtAIMOErEpWBvtyMYbEswTKWN4gQGnBuGqHjaEQy25ZQwmZanBNpbHBLiUvi0IxhsSx5NpuWRgNLiRBCXwrQjGGwr+HWK4vyKw9f/EBZVRkwt879z/wEAAP//AAAA//+yKUhMT/VNLErPzCtWyElNK7FVMtAzV1IoykzPgLFL8gvAoqZKCkn5JSX5uTBeRmpiSmoRiGespJCWn18C4+jb2eiX5xdlF2ekppbYAQAAAP//AwBQSwMEFAAGAAgAAAAhAPZgtEG4BwAAESIAABMAAAB4bC90aGVtZS90aGVtZTEueG1s7FrNjxu3Fb8HyP9AzF3WzOh7YTnQpzf27nrhlV3kSEmUhl7OcEBSuysUAQrn1EuBAmnRS4HeeiiKBmiABrnkjzFgI03/iDxyRprhioq9/kCSYncvM9TvPf7mvcfHN49z95OrmKELIiTlSdcL7vgeIsmMz2my7HpPJuNK20NS4WSOGU9I11sT6X1y7+OP7uIDFZGYIJBP5AHuepFS6UG1KmcwjOUdnpIEfltwEWMFt2JZnQt8CXpjVg19v1mNMU08lOAY1D5aLOiMoIlW6d3bKB8xuE2U1AMzJs60amJJGOz8PNAIuZYDJtAFZl0P5pnzywm5Uh5iWCr4oev55s+r3rtbxQe5EFN7ZEtyY/OXy+UC8/PQzCmW0+2k/ihs14OtfgNgahc3auv/rT4DwLMZPGnGpawzaDT9dphjS6Ds0qG70wpqNr6kv7bDOeg0+2Hd0m9Amf767jOOO6Nhw8IbUIZv7OB7ftjv1Cy8AWX45g6+Puq1wpGFN6CI0eR8F91stdvNHL2FLDg7dMI7zabfGubwAgXRsI0uPcWCJ2pfrMX4GRdjAGggw4omSK1TssAziOJeqrhEQypThtceSnHCJQz7YRBA6NX9cPtvLI4PCC5Ja17ARO4MaT5IzgRNVdd7AFq9EuTlN9+8eP71i+f/efHFFy+e/wsd0WWkMlWW3CFOlmW5H/7+x//99Xfov//+2w9f/smNl2X8q3/+/tW33/2UelhqhSle/vmrV19/9fIvf/j+H186tPcEnpbhExoTiU7IJXrMY3hAYwqbP5mKm0lMIkwtCRyBbofqkYos4MkaMxeuT2wTPhWQZVzA+6tnFtezSKwUdcz8MIot4DHnrM+F0wAP9VwlC09WydI9uViVcY8xvnDNPcCJ5eDRKoX0Sl0qBxGxaJ4ynCi8JAlRSP/GzwlxPN1nlFp2PaYzwSVfKPQZRX1MnSaZ0KkVSIXQIY3BL2sXQXC1ZZvjp6jPmeuph+TCRsKywMxBfkKYZcb7eKVw7FI5wTErG/wIq8hF8mwtZmXcSCrw9JIwjkZzIqVL5pGA5y05/SGGxOZ0+zFbxzZSKHru0nmEOS8jh/x8EOE4dXKmSVTGfirPIUQxOuXKBT/m9grR9+AHnOx191NKLHe/PhE8gQRXplQEiP5lJRy+vE+4vR7XbIGJK8v0RGxl156gzujor5ZWaB8RwvAlnhOCnnzqYNDnqWXzgvSDCLLKIXEF1gNsx6q+T4iEMknXNbsp8ohKK2TPyJLv4XO8vpZ41jiJsdin+QS8boXuVMBidFB4xGbnZeAJhfIP4sVplEcSdJSCe7RP62mErb1L30t3vK6F5b83WWOwLp/ddF2CDLmxDCT2N7bNBDNrgiJgJpiiI1e6BRHL/YWI3leN2Mopt7AXbeEGKIyseiemyeuKnxMsBL/8eWqfD1b1uBW/S72zL68cXqty9uF+hbXNEK+SUwLbyW7iui1tbksb7/++tNm3lm8LmtuC5ragcb2CfZCCpqhhoLwpWj2m8RPv7fssKGNnas3IkTStHwmvNfMxDJqelGlMbvuAaQSX+nlgAgu3FNjIIMHVb6iKziKcQn8oMF3MpcxVLyVKuYS2kRk2/VRyTbdpPq3iYz7P2p2mv+RnJpRYFeN+AxpP2Ti0qlSGbrbyQc1vQ92wXZpW64aAlr0JidJkNomag0RrM/gaErpz9n5YdBws2lr9xlU7pgBqW6/AezeCt/Wu16hnjKAjBzX6XPspc/XGu9o579XT+4zJyhEArcVdT3c0172Pp58uC7U38LRFwjglCyubhPGVKfBkBG/DeXSW++4/FXA39XWncKlFT5tisxoKGq32h/C1TiLXcgNLypmCJegS1ngIi85DM5x2vQX0jeEyTiF4pH73wmwJhy8zJbIV/zapJRVSDbGMMoubrJP5J6aKCMRo3PX082/DgSUmiWTkOrB0f6nkQr3gfmnkwOu2l8liQWaq7PfSiLZ0dgspPksWzl+N+NuDtSRfgbvPovklmrKVeIwhxBqtQHt3TiUcHwSZq+cUzsO2mayIv2s7U579rUOuIh9jlkY431LK2TyDmw1lS8fcbW1QusufGQy6a8LpUu+w77ztvn6v1pYr9sdOsWlaaUVvm+5s+uF2+RKrYhe1WGW5+3rO7WySHQSqc5t4972/RK2YzKKmGe/mYZ2081Gb2nusCEq7T3OP3babhNMSb7v1g9z1qNU7xKawNIFvDs7LZ9t8+gySxxBOEVcsO+1mCdyZ0jI9Fca3Uz5f55dMZokm87kuSrNU/pgsEJ1fdb3QVTnmh8d5NcASQJuaF1bYVtBZ7dmCerPLRbMFuxXOythr9aotvJXYHLNuhU1r0UVbXW1O1HWtbmbWDsue2qRhYym42rUitMkFhtI5O8zNci/kmSuVV9pwhVaCdr3f+o1efRA2BhW/3RhV6rW6X2k3erVKr9GoBaNG4A/74edAT0Vx0Mi+fBjDaRBb598/mPGdbyDizYHXnRmPq9x841A13jffQATh/m8gwJFAKxwF9bAXDiqDYdCs1MNhs9Ju1XqVQdgchj3YtJvj3uceujDgoD8cjseNsNIcAK7u9xqVXr82qDTbo344Dkb1oQ/gfPu5grcYnXNzW8Cl4XXvRwAAAP//AwBQSwMEFAAGAAgAAAAhAHh4ZNSsBgAAVTgAAA0AAAB4bC9zdHlsZXMueG1s7Ftbb+I4FH5faf9DlHeaCwQIAkbTdpBGmq0qTVfa1xAMtSaJkWM6MKv973Ps3AwhkKQQUmmRqiZOfHyun49P7PGnre8pb4iGmAQT1bjTVQUFLlngYDVR/36ZdYaqEjInWDgeCdBE3aFQ/TT9849xyHYe+v6KEFOARBBO1FfG1iNNC91X5DvhHVmjAJ4sCfUdBrd0pYVripxFyDv5nmbqel/zHRyoEYWR75Yh4jv0x2bdcYm/dhieYw+znaClKr47+roKCHXmHrC6NXqOq2yNPjWVLU0GEa25cXzsUhKSJbsDuhpZLrGL8uzamq05bkYJKNejZFiabu7JvqU1KfU0it4wN586HS9JwELFJZuATVSwnBB29CMgP4MZfwQWjt+ajsNfypvjQYuhatOxSzxCFQamA82JlsDxUfTG5zUjofLkUEp+8neXjo+9XfTM5A3C5PHLPgYD8EaNMxOx1PBgc85SsXQ65+6i0p0ZsBl14mCBtmjB7d6UfHQ1n6iz+Fdx1LPuERPXxe89xIusc4EBhBAhODn2vDTuIKCilukYEIohGszgsRJfv+zWEGEBgGkUJPDo7Nsr6uwM0yrfISQeXnAuVg8irlNDAeT2OZl5/CB1mn5PUJcY5hFchrnDsWIM6aoKwxyG9DuzZ9uDgcF/g6HdvfL4qawP4DnNyDpoZph+ptKuDSq1rKFl2GYP/gTEXN+oMKUkRr0RB4OEg07TfpX36yEYwdaHPfgNDGtgNB9ZA4gs2+wO9O7A7HcrO4EIcICvOaELSAGTxMEwATqitunYQ0sGgEHx6pX/Z2TN4YMwBnnSdLzAzooEjsfn+6SH3BNyR0gTJyp7hTQvyTIOQYcPEY9Q6n3Bi2Cl1OvAcsJxqfcj4W4u2zXVfGDIi6olcZaKpkzF/ZA2+khM1w0aHy3wxi+M4paETVXpSoBYTY++IidR/N4U/RpSylWg6dLwF099MJO6yPO+8ynvn2U6nfLZdLtUgo0/89lXWKBBjYWvjZNLSLLjy2jmjG6Ax71OdtbJgMvjnRRnvfZ2fKkvRonu4NXs7l5M9dn9Zw+vAh/JHZ4pYchloiAkEmhNliqSURLPsmvJp2yXxwU9rp2eqmTaAQUkvWOJhYzQLCROZBT5UJHmzUyJkNVmtCHJPqAttJnQ5ks5XiA5HCm6l7TJF4FOolzllVD8CwjxgosL2kZQj4KqG8Ou1MKTqO2y2FmKWM6ro/Usgyit0rIUXdz/4uiC1swxrITl1BVao2WJZbg8ESdZZFTwV56e1/DWIqZgBbdv+iaZkkIILo8YF7iWMPRmUc/5PDtFHOjxaePPEZ2Jentcw90DP4lmEZLkwvKQ5kdRThqsiT2PC1LL9arhdwlL5iLiZsxKcy/U74qSjOPIfZiwlMhycla6wOx6YbQqnWrcDlDzWrz0xHR1nbaEYcj9jvr8mQSrNIpUzvjAtAlDe9PV9TO+Q5NDEU18IDiVnzbLbWVlAsxe1bqnc/wyCqyRx51A3SIEv5Q3H5H33ApGErDegu5dy6yKJmiKwwrLvoZBoHamA6o7vYY6nZZFBYw2JN9SDBmnl1ZXzJRzSKf8pM76BW2hyiNqHPmqgZR1FopQMke+iDHeJ0KRO+WQrO3LlSJBSq4A2hMXpWeqmy1lSlR0Uvf5QBWdnMsXFzuPrzw/KoxJHrcvQpOxUwPGpGXwtZcX70vI9qa3yxTpale6i+AlV5/6v4Cz/5WoRLWpZI2vAc1KoVFUl23xp40iH82lVVU/RwFMnPrU1fqvDjDzFhX4a9QIOdjXICc5FGDZh9bnCdA7oc+iiLp+tfAS3zfbstA9v9IChEq+1Egqh0XXCZ+79TLlvFTHoflMRaQxZKrEfxFO8+MR8RaECtsGiqiBwS9IDRT9Pmp7lsolc7fcxFBUAsyhXEt2WgBetmr2EJtxYPuNtNdob6dRuldH4UdcJuoT/x7sSf4032APNnGnm28OOzwjyhFc2u0g9YjOIKRdgIvFNtvpJHYLMX76SuyBSvkCmy/Q0tl47CV9OFGz67/EzkLwgPitZ/xGmCAxUbPrb3wLcrTJGkpf30LYMwz/lQ3FE/XfL/cD+/HLzOwM9fthp9dFVse27h87Vu/h/vFxZuum/vCfdAbsHSfAxJE12KVj9EahB+fEaCxszPz3rG2iSjcR++IjCrAt826bff2zZeidWVc3Or2+M+wM+12rM7MM87Hfu/9izSyJd6vmSTFdM4zozBln3hox7CMPB4mtEgvJrWAkuD0hhJZYQsvOA05/AwAA//8DAFBLAwQUAAYACAAAACEAtMZpgToEAACcDQAAFAAAAHhsL3NoYXJlZFN0cmluZ3MueG1srFdtb9s2EP5eoP/hoGGDB9hRkr6uc1w4stMZs2M1UrrtIyedLSESqZGUa//7HWXLCSTacbd+COAceQ+Px7vnHvU/rvMMVihVKviVc3F27gDySMQpX1459+FN770DSjMes0xwvHI2qJyPg5cv+kppIF+urpxE6+KD66oowZypM1Egp5WFkDnT9K9cuqqQyGKVIOo8cy/Pz9+6OUu5A5EouaZzz185UPL0nxK9reXde2fQV+mgrwce07gUctN39aDvGtvWHgx9GM4CGC/oKK2gMyyKLI2YpqvAFFeY/dxwMQF/UAWL6CIUkUK5QmfwRWRljtCEH84hKIuCoOHlix9bh6d5kWHTOsM4LfOm1RNm77ppDoVmWR19x//zczPcwfge7rym263gVvtNOG4dPA/CoGn07ybeuGWtk5hySKRq+tCBFvhtJJaFORUUyzKbDyF5QunWCVss61KNZl0kPF+mER4AtK/ViPbVmeCpFpI6oJmGH0AswJTdNVOpooNFXEZVuQUbpTFXQOnTCcKUGkZFrED4muoEflm/eQaKrt8zcPB/gU6O6eLt9wrqBKSTo7p8vX73fVJlQwol48p0dKta9i+73wJ3SGxElQo920MXKCEXXCcHo7Ui0Tt/I94knEyfq8ML17+ECY/SGDkF/Hxsu/p95fqvXf/Nf3D1EsaXePpRwd3nb4hLir8z000HL/IXMpltYEjNt0p12u7+/XuaMaPpj/EIgZn9ZvsJKdo2IlXBggZFQs1OD76hYw++RT0sfBY9sDo3Rz1+R8kxg/tiKVm8C+qow/bZ6oOIayZcEwZNERfG691PVsbprghsYJ+QG3KGkKmHVhvMUhV1PMEXKEkFIHjE4qoLIxHRhKQ8Gq5rTSlDiJ8EUX6V5IP5mfuzetgdzqHXO2WbN5q5weh5OC84uGdc0tMqUcqoNcQNFx9ZDs7gVrTkAGmIahAMlSrzwvxsJXeEVIhZyzzkGxA0NCSoiKQT0GCJyJt6WWIMq5QBg2271ZQEbFvAVM4lawOGvXqIqMRM4VJhNZSiUmmRk9QDLUSmgFRTNcyIxnbzDn5iefErTMJg1rzgkXsFVdg0Gf8Q8qHhZ1dc1QgdBpNgr7Ba+TRDttpR01q3TgKNVkPIhiEgZ5xazZTmyec+0vKjc0v7VdLS9NE+wIUU+S4iYiVVoKGSVunMS11JBJORUyOi9rruwjyAKckYrogHbMl4qm1HRtuKwlybRO/I2pBP98/2Wdo9cM0gHZLPVu+QBh/pH+jclLwSOCzrwlSwuGt4sFL2RA/din+IvKrK39WOYaQFKWwr7iNcndjmXU3zVdl/ouV34VrTcj303RCjhJPwzw6BGsAv22+cPdva0CZ8IZnSkkRdKWlcUKXdov5KZX0M+TfMihGqByLhgL4o6BWh+vdA2KRYW18Tlfp9YnXp62rwLwAAAP//AwBQSwMEFAAGAAgAAAAhADttMkvBAAAAQgEAACMAAAB4bC93b3Jrc2hlZXRzL19yZWxzL3NoZWV0MS54bWwucmVsc4SPwYrCMBRF9wP+Q3h7k9aFDENTNyK4VecDYvraBtuXkPcU/XuzHGXA5eVwz+U2m/s8qRtmDpEs1LoCheRjF2iw8HvaLb9BsTjq3BQJLTyQYdMuvpoDTk5KiceQWBULsYVRJP0Yw37E2bGOCamQPubZSYl5MMn5ixvQrKpqbfJfB7QvTrXvLOR9V4M6PVJZ/uyOfR88bqO/zkjyz4RJOZBgPqJIOchF7fKAYkHrd/aea30OBKZtzMvz9gkAAP//AwBQSwMEFAAGAAgAAAAhAM/gR7TCAQAALBUAACcAAAB4bC9wcmludGVyU2V0dGluZ3MvcHJpbnRlclNldHRpbmdzMS5iaW7sVM1K3FAYPTOx7eimDhTcdFGkK3HoDJOp3VWZpHZK0oQkM7hxMXRSCIzJkEREpYL4Gj5Ily5d9gG6diHFB3Cj56Yz2JahjOBG+O7lu9/PPTk395B8NiJ8QYoEGe0rcryCyzxCXMQ5q6pi4AOmjdKc9vQn3BfamxLUvFxIKgP659gql+m3yhpXCyHZcq7pVJb7FUtjuPJlmvI3HJsdX/+Tyeh87i7jHK+11er77cOj/53ypNicL7ge4BWF4hEqMPmuZnn1c4J8O/iksIv4jkPU8Q46/5I6Glw3UIOJt2iyVqMZWOOsEdNk3WRUZ64zb9C3mTXRKrJvZPRM37AsdOMoDTMVuf1RmPrRQQjLDALTg5NGYZz38yiJ4Tpe4G10Anhhlgx3ixpDZ6SiBtrJMEntZBD+jv6+3WoV6OmGPbn76cJo+SUhv2ga7brkVPSLPfvk6tnHpbPW8Q/WrPEeKndcCqvylbFX+Tqtp/JF8P4J+8wudtgDVGfpst+obuCizyjDHvdTDAj+F+lwL54R2ybHPkbk9/mEOk91spw1GaKAKCAKiAKigCggCogCooAoIAqIAqKAKCAKzKLALQAAAP//AwBQSwMEFAAGAAgAAAAhAEBUkGDJAAAAKwEAABkAAABkb2NNZXRhZGF0YS9MYWJlbEluZm8ueG1slM9LasMwEIDhq4jZy1KMHIKxnHUhPYQeo1qgR7CmIaH07pV37bK7YWA+/lmuz5zYA/cWa9FwGiQwLK76WD40fFLgF2CNTPEm1YIaXtjgui4u2TQnYzHdYiPWkdLmY6lhI7rPQjS3YTZtyNHttdVAg6tZ1BCiQzHKUYoc77dDeEcy3pCB3yyLXsOX9ah8CI4re5ZcuXDmFysVNxOiHacwORW+j2JjE/aDXp+RttrHnh0J3/6h7Jjr41BOwMS6iL8/rj8AAAD//wMAUEsDBBQABgAIAAAAIQArJoB+VQEAAH8CAAARAAgBZG9jUHJvcHMvY29yZS54bWwgogQBKKAAAQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAACEklFLwzAUhd8F/0PJk4Jt2q6bEtoOVPbkYGBF8S0kd1tYm4Yk2u3fm7ZbrUzxMTnnfjnnknS+r0rvE7QRtcxQFITIA8lqLuQmQy/Fwr9DnrFUclrWEjJ0AIPm+eVFyhRhtYaVrhVoK8B4jiQNYSpDW2sVwdiwLVTUBM4hnbiudUWtO+oNVpTt6AZwHIYzXIGlnFqKW6CvBiI6IjkbkOpDlx2AMwwlVCCtwVEQ4W+vBV2ZXwc6ZeSshD0o1+kYd8zmrBcH996Iwdg0TdBMuhguf4Tflk/PXVVfyHZXDFCeckaYBmprnS9hR0t64622VArJNfWuYG+vUzzytPssqbFLt/q1AH5/+Hvs3Ope68r1TwL3XFzSlzspr5OHx2KB8jiMp3448+NpEYUkikgSv7dJfsy38fuL6pjnX+KtH8ZFmJBpQuJkRDwB8hSffZn8CwAA//8DAFBLAwQUAAYACAAAACEAdi2D2Z4BAABLAwAAEAAIAWRvY1Byb3BzL2FwcC54bWwgogQBKKAAAQAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAAACck1Fr2zAQx98H+w5G743crJQRZJWStvRhY4G47WO5yedYVJaMdDHJPv0kmzrO+lDY2939j79+upPEzaE1WY8+aGcLdrnIWYZWuUrbXcGeyoeL7ywLBLYC4ywW7IiB3civX8TGuw49aQxZtLChYA1Rt+I8qAZbCIso26jUzrdAMfU77upaK7xzat+iJb7M82uOB0JbYXXRTYZsdFz19L+mlVOJLzyXxy4CS1E6AlPqFmUu+CkRt11ntAKKt5c/tfIuuJqy+4NCI/hcFJF6i2rvNR2TxzwVWwUG1/FAWYMJKPipIB4R0jA3oH2QoqdVj4qcz4L+E8e5ZNlvCJgwC9aD12Ap4qa2MRli0wXy8sX5t9AgUhA8NozFIZz3zmN9JZdDQwzOG5PBCBKFc8RSk8Hwq96Ap8+IB4aRd8SptQXzOlDOISfcrTP7NOwPNxiGEln+OX3t2g7sMQpT9EPbt/DUle4OCN8Hfl4U2wY8VnFH00KmgniMs/YmmawbsDus3ns+Cul5PI9/Q15eL/Jvedz8rCb46RfIvwAAAP//AwBQSwECLQAUAAYACAAAACEATMqDoJUBAAD+BQAAEwAAAAAAAAAAAAAAAAAAAAAAW0NvbnRlbnRfVHlwZXNdLnhtbFBLAQItABQABgAIAAAAIQAxHYnNIgEAAN4CAAALAAAAAAAAAAAAAAAAAM4DAABfcmVscy8ucmVsc1BLAQItABQABgAIAAAAIQCSJ1TfxQMAAEkJAAAPAAAAAAAAAAAAAAAAACEHAAB4bC93b3JrYm9vay54bWxQSwECLQAUAAYACAAAACEASqmmYfoAAABHAwAAGgAAAAAAAAAAAAAAAAATCwAAeGwvX3JlbHMvd29ya2Jvb2sueG1sLnJlbHNQSwECLQAUAAYACAAAACEApNfiL1wMAADOPwAAGAAAAAAAAAAAAAAAAABNDQAAeGwvd29ya3NoZWV0cy9zaGVldDEueG1sUEsBAi0AFAAGAAgAAAAhACiZDrOjBQAAjRgAABgAAAAAAAAAAAAAAAAA3xkAAHhsL3dvcmtzaGVldHMvc2hlZXQyLnhtbFBLAQItABQABgAIAAAAIQD2YLRBuAcAABEiAAATAAAAAAAAAAAAAAAAALgfAAB4bC90aGVtZS90aGVtZTEueG1sUEsBAi0AFAAGAAgAAAAhAHh4ZNSsBgAAVTgAAA0AAAAAAAAAAAAAAAAAoScAAHhsL3N0eWxlcy54bWxQSwECLQAUAAYACAAAACEAtMZpgToEAACcDQAAFAAAAAAAAAAAAAAAAAB4LgAAeGwvc2hhcmVkU3RyaW5ncy54bWxQSwECLQAUAAYACAAAACEAO20yS8EAAABCAQAAIwAAAAAAAAAAAAAAAADkMgAAeGwvd29ya3NoZWV0cy9fcmVscy9zaGVldDEueG1sLnJlbHNQSwECLQAUAAYACAAAACEAz+BHtMIBAAAsFQAAJwAAAAAAAAAAAAAAAADmMwAAeGwvcHJpbnRlclNldHRpbmdzL3ByaW50ZXJTZXR0aW5nczEuYmluUEsBAi0AFAAGAAgAAAAhAEBUkGDJAAAAKwEAABkAAAAAAAAAAAAAAAAA7TUAAGRvY01ldGFkYXRhL0xhYmVsSW5mby54bWxQSwECLQAUAAYACAAAACEAKyaAflUBAAB/AgAAEQAAAAAAAAAAAAAAAADtNgAAZG9jUHJvcHMvY29yZS54bWxQSwECLQAUAAYACAAAACEAdi2D2Z4BAABLAwAAEAAAAAAAAAAAAAAAAAB5OQAAZG9jUHJvcHMvYXBwLnhtbFBLBQYAAAAADgAOALMDAABNPAAAAAA='''

        # Decode it and load into openpyxl
        decoded_excel = base64.b64decode(encoded_excel_string)
        wb = load_workbook(io.BytesIO(decoded_excel))
        ws = wb.active

        try:
            
            # Copying volume into Solution sheet
            assumption_Sheet = wb['Solution']
            # Write volume values from self.volume[0–17] to column D, rows 4–21
            for i in range(18):
                assumption_Sheet.cell(row=4 + i, column=4).value = self.volume[i]  # column=4 is 'D'


            #  Map: DataFrame column name -> Excel column letter
            self.column_map = {
                                    'volume':'D',
                                    'total_efforts': 'J',
                                    'eu' : 'E',
                                    'non_eu' : 'F',
                                    'dummy_complexity_matrix' : temp,

                                    'eu_resource' : 'K',
                                    'eu_rate_card': 'L',
                                    'non_eu_resource': 'M',
                                    'non_eu_rate_card': 'N',
                                    
                                    'eu_fte': 'O',
                                    'non_eu_fte': 'P',
                                    'overall_fte': 'Q',

                                    'eu_cost': 'R',
                                    "non_eu_cost": 'S',
                                    'overall_cost': 'T',

                                    'eu_price': 'U',
                                    'non_eu_price': 'V',
                                    'overall_price': 'W'
                        }
            
            start_row = 5
            for i in range(23):
                ws[f"{self.column_map['volume']}{i + start_row}"] = self.volume[i]
                ws[f"{self.column_map['eu']}{i + start_row}"] = self.eu[i] 
                ws[f"{self.column_map['non_eu']}{i + start_row}"] = self.non_eu[i] 
                ws[f"{self.column_map['dummy_complexity_matrix']}{i + start_row}"] = self.complexity_matrix[i][self.mode]
                ws[f"{self.column_map['total_efforts']}{i + start_row}"] = self.total_efforts[i]

                ws[f"{self.column_map['eu_resource']}{i + start_row}"] = self.eu_resource[i]
                ws[f"{self.column_map['eu_rate_card']}{i + start_row}"] = self.eu_rate_card[i]
                ws[f"{self.column_map['non_eu_resource']}{i + start_row}"] = self.non_eu_resource[i]
                ws[f"{self.column_map['non_eu_rate_card']}{i + start_row}"] = self.non_eu_rate_card[i]

                ws[f"{self.column_map['eu_fte']}{i + start_row}"] = round(self.eu_fte[i], 2)
                ws[f"{self.column_map['non_eu_fte']}{i + start_row}"] = round(self.non_eu_fte[i], 2)
                ws[f"{self.column_map['overall_fte']}{i + start_row}"] = round(self.overall_fte[i], 2)

                ws[f"{self.column_map['eu_cost']}{i + start_row}"] = round(self.eu_cost[i], 2)
                ws[f"{self.column_map['non_eu_cost']}{i + start_row}"] = round(self.non_eu_cost[i], 2)
                ws[f"{self.column_map['overall_cost']}{i + start_row}"] = round(self.overall_cost[i], 2)

                ws[f"{self.column_map['eu_price']}{i + start_row}"] = round(self.eu_price[i], 2)
                ws[f"{self.column_map['non_eu_price']}{i + start_row}"] = round(self.non_eu_price[i], 2)
                ws[f"{self.column_map['overall_price']}{i + start_row}"] = round(self.overall_price[i], 2)

            # These columns can be summed (all numeric fields)
            summable_columns = [
                                    'eu_fte', 'non_eu_fte', 'overall_fte',
                                    'eu_cost', 'non_eu_cost', 'overall_cost',
                                    'eu_price', 'non_eu_price', 'overall_price'
                               ]

            sum_row = start_row + 23
            ws[f"B{sum_row}"] = "TOTAL"

            # sum 
            for key in summable_columns:
                if key in self.column_map:
                    col_letter = self.column_map[key]
                    total = round(sum(getattr(self, key)), 2)  # round sum of Python list
                    ws[f"{col_letter}{sum_row}"] = total

            # First unhide all three complexity columns
            for col in ['G', 'H', 'I']:
                ws.column_dimensions[col].hidden = False

            # Then hide only the two that are not selected
            complexity_map = {'G': 0, 'H': 1, 'I': 2}
            for col_letter, complexity_value in complexity_map.items():
                if self.mode != complexity_value:
                    ws.column_dimensions[col_letter].hidden = True


            # Hide only columns in Excel (UI stays unchanged)
            sheet_main = wb.active

            # First, unhide all resource columns
            for col in ['K', 'L', 'M', 'N']:
                sheet_main.column_dimensions[col].hidden = False

            if region_mode == "eu":
                sheet_main.column_dimensions['M'].hidden = True  # Non-EU Resource
                sheet_main.column_dimensions['N'].hidden = True  # Non-EU Rate

            elif region_mode == "non-eu":
                sheet_main.column_dimensions['K'].hidden = True  # EU Resource
                sheet_main.column_dimensions['L'].hidden = True  # EU Rate


            # First, unhide all EU/Non-EU effort, cost, and price columns
            for col in ['E', 'F', 'O', 'P', 'R', 'S', 'U', 'V']:
                sheet_main.column_dimensions[col].hidden = False

            # Now hide based on region selection
            if region_mode == "eu":
                for col in ['F', 'P', 'S', 'V']:  # Non-EU related columns
                    sheet_main.column_dimensions[col].hidden = True

            elif region_mode == "non-eu":
                for col in ['E', 'O', 'R', 'U']:  # EU related columns
                    sheet_main.column_dimensions[col].hidden = True

            # Hide overall summary columns if region is not "both"
            for col in ['Q', 'T', 'W']:
                sheet_main.column_dimensions[col].hidden = (region_mode != 'both')


            wb.save(output_path)
            print(f"\n Excel file saved successfully as: {output_path}\n")
            return os.path.abspath(output_path)

        except Exception as e:
            print(f"\n Error: {e}\n")
            return None
    
'''
sap = SAP()
sap.user_inputs()
sap.calculate_efforts()
sap.final_calculations()
sap.fill_excel()
'''



import tkinter as tk
from tkinter import ttk, messagebox

class SAP_GUI:
    def __init__(self, root):
        self.root = root
        self.root.title("SAP AMS Estimation Tool")
        self.sap = SAP()

        self.inputs = []
        self.complexity_entries = []
        self.eu_dropdowns = []
        self.non_eu_dropdowns = []
        self.eu_labels = []
        self.non_eu_labels = []
        self.eu_rate_labels = []
        self.non_eu_rate_labels = []

        self.user_edit_color = "#add8e6"   # Light blue
        self.auto_fill_color = "#90ee90"   # Light green
        
        self.mode = tk.IntVar(value=0)  # Set default complexity mode to Simple

        self.default_efforts = []
        self.default_roles_eu = []
        self.default_roles_noneu = []
        self.default_rate_eu = []
        self.default_rate_noneu = []

        self.build_gui()
        self.fill_default_complexity()


    def build_gui(self):

        style = ttk.Style()
        style.configure("Green.TCombobox", fieldbackground="#90ee90")
        style.configure("Blue.TCombobox", fieldbackground="#add8e6")

        # ---- Group Customer Name + Complexity in One Row ----
        top_frame = tk.Frame(self.root, bg="#f0f8ff")
        top_frame.grid(row=0, column=0, columnspan=10, sticky='w', pady=10)

        # Customer Name (Left Side)
        tk.Label(top_frame, text="Customer Name:", font=("Arial", 11, "bold"), bg="#f0f8ff").pack(side=tk.LEFT, padx=(5, 2))
        self.name_entry = tk.Entry(top_frame, font=("Arial", 10), width=30)
        self.name_entry.pack(side=tk.LEFT, padx=(0, 80))
        self.name_entry.focus()
        
        # Complexity Selection (Right Side)
        tk.Label(top_frame, text="Select Complexity:", font=("Arial", 11, "bold"), bg="#f0f8ff").pack(side=tk.LEFT, padx=(5, 5))
        self.mode = tk.IntVar()
        for i, label in enumerate(['Simple', 'Medium', 'Complex']):
            tk.Radiobutton(top_frame, text=label, variable=self.mode, value=i, bg="#f0f8ff", command=self.fill_default_complexity).pack(side=tk.LEFT, padx=2)
        

        # --- EU/Non-EU Selection ---
        tk.Label(top_frame, text="Select Region:", font=("Arial", 11, "bold"), bg="#f0f8ff").pack(side=tk.LEFT, padx=(20, 5))
        self.region_mode = tk.StringVar(value="both")  # default
        for value in ["eu", "non-eu", "both"]:
            tk.Radiobutton(top_frame, text=value.upper(), variable=self.region_mode, value=value, bg="#f0f8ff", command=self.update_region_view).pack(side=tk.LEFT, padx=2)


        # Column headers
        headers = ["Input", "Volume", "Efforts in hrs", "EU %", "Non-EU %", "EU Resource", "Non-EU Resource", "EU Rate Card", "Non-EU Rate Card"]
        for col, header in enumerate(headers):
            tk.Label(self.root, text=header, font=('Arial', 10, 'bold')).grid(row=2, column=col, padx=5, pady=5)


        # Input rows
        for i, col_name in enumerate(self.sap.input_cols):
            # Uniform font and color for all rows
            tk.Label(self.root, text=col_name, anchor='w', justify='left',
                    font=('Arial', 10), fg='black').grid(row=i+3, column=0, sticky='w', padx=5, pady=2)

            # Volume entry (skip for FTE rows)
            if i < 19:
                vol_entry = tk.Entry(self.root, width=5)
                vol_entry.grid(row=i+3, column=1)
                vol_entry.config(bg=self.user_edit_color)
                vol_entry.bind("<Return>", lambda e, index=i: self.focus_next(self.inputs, index))
                vol_entry.bind("<FocusOut>", self.on_user_edit)
                vol_entry.bind("<KeyRelease>", self.on_user_edit)
                self.inputs.append(vol_entry)
            else:
                self.inputs.append(None)

            complexity_entry = tk.Entry(self.root, width=5)
            complexity_entry.grid(row=i+3, column=2)
            complexity_entry.config(bg=self.auto_fill_color)
            complexity_entry.bind("<Return>", lambda e, index=i: self.focus_next(self.complexity_entries, index))
            complexity_entry.bind("<FocusOut>", self.on_user_edit)
            complexity_entry.bind("<KeyRelease>", self.on_user_edit)
            self.complexity_entries.append(complexity_entry)


            # --- EU % Entry ---
            eu_entry = tk.Entry(self.root, width=5)
            eu_entry.insert(0, "50")  # Default EU %
            eu_entry.grid(row=i+3, column=3)
            eu_entry.config(bg=self.user_edit_color)
            eu_entry.bind("<Return>", lambda e, index=i: self.focus_next(self.eu_dropdowns, index))
            self.eu_dropdowns.append(eu_entry)
            eu_entry.bind("<FocusOut>", self.on_user_edit)
            eu_entry.bind("<KeyRelease>", self.on_user_edit)

            # --- Non-EU % Entry ---
            non_eu_entry = tk.Entry(self.root, width=5)
            non_eu_entry.insert(0, "50")  # Default Non-EU %
            non_eu_entry.grid(row=i+3, column=4)
            non_eu_entry.config(bg=self.user_edit_color)
            non_eu_entry.bind("<Return>", lambda e, index=i: self.focus_next(self.non_eu_dropdowns, index))
            self.non_eu_dropdowns.append(non_eu_entry)
            non_eu_entry.bind("<FocusOut>", self.on_user_edit)
            non_eu_entry.bind("<KeyRelease>", self.on_user_edit)

            # --- Resource Type Options ---
            options = ["Junior", "Senior", "Specialist", "Expert"]



            # --- Set Default Resource Role Based on Row Index ---
            if i < 8:
                default_role = "Junior"
            elif i < 19:
                default_role = "Specialist"
            else:
                default_role = "Expert"

            # --- EU Resource Combobox ---
            eu_cb = ttk.Combobox(self.root, values=options, width=10, state="readonly")
            eu_cb.grid(row=i+3, column=5)
            eu_cb.set(default_role)
            eu_cb.configure(background=self.auto_fill_color)
            eu_cb.bind("<<ComboboxSelected>>", lambda e, index=i: self.on_resource_change(index, 'eu'))
            self.eu_labels.append(eu_cb)

            # --- Non-EU Resource Combobox ---
            non_eu_cb = ttk.Combobox(self.root, values=options, width=10, state="readonly")
            non_eu_cb.grid(row=i+3, column=6)
            non_eu_cb.set(default_role)
            non_eu_cb.configure(background=self.auto_fill_color)
            non_eu_cb.bind("<<ComboboxSelected>>", lambda e, index=i: self.on_resource_change(index, 'non_eu'))
            self.non_eu_labels.append(non_eu_cb)


            # EU Rate Card Label and non-eu label
            eu_rate_entry = tk.Entry(self.root, width=5)
            eu_rate_entry.insert(0, "0")
            eu_rate_entry.config(bg=self.auto_fill_color)
            eu_rate_entry.bind("<FocusOut>", self.on_user_edit)
            eu_rate_entry.bind("<KeyRelease>", self.on_user_edit)
            eu_rate_entry.grid(row=i+3, column=7)
            self.eu_rate_labels.append(eu_rate_entry)

            non_eu_rate_entry = tk.Entry(self.root, width=5)
            non_eu_rate_entry.insert(0, "0")
            non_eu_rate_entry.insert(0, "0")
            non_eu_rate_entry.config(bg=self.auto_fill_color)
            non_eu_rate_entry.bind("<FocusOut>", self.on_user_edit)
            non_eu_rate_entry.bind("<KeyRelease>", self.on_user_edit)
            non_eu_rate_entry.grid(row=i+3, column=8)
            self.non_eu_rate_labels.append(non_eu_rate_entry)

            # --- Now it's safe to auto-fill rate cards ---
            self.update_rate_card(i, 'eu')
            self.update_rate_card(i, 'non_eu')

            self.default_roles_eu.append(default_role)
            self.default_roles_noneu.append(default_role)

            self.default_rate_eu.append(str(self.eu_rate_labels[i].get()))
            self.default_rate_noneu.append(str(self.non_eu_rate_labels[i].get()))
            


        # Buttons
        btn_frame = tk.Frame(self.root)
        btn_frame.grid(row=30, column=0, columnspan=9, pady=20)

        tk.Button(btn_frame, text="Preview Summary", command=self.show_summary, width=20).pack(side='left', padx=10)
        tk.Button(btn_frame, text="Export to Excel", command=self.export_excel, width=20).pack(side='left', padx=10)
        tk.Button(btn_frame, text="Exit", command=self.root.quit, width=10).pack(side='left', padx=10)

        # --- Color Legend Info ---
        legend_frame = tk.Frame(self.root, bg="#f0f8ff")
        legend_frame.grid(row=31, column=0, columnspan=9, sticky='w', padx=10, pady=(0, 10))

        # --- Color Legend Info (Single Line) ---
        legend_frame = tk.Frame(self.root, bg="#f0f8ff")
        legend_frame.grid(row=31, column=0, columnspan=9, sticky='w', padx=10, pady=(0, 10))

        tk.Label(legend_frame, text="NOTE :", font=("Arial", 10, "bold"), bg="#f0f8ff").grid(row=0, column=0, sticky='w')
        tk.Label(legend_frame, text="Green = Default Value (Adjustable)", bg="#90ee90", font=("Arial", 10), width=30).grid(row=0, column=1, padx=10)
        tk.Label(legend_frame, text="Blue = User Edited", bg="#add8e6", font=("Arial", 10), width=25).grid(row=0, column=2, padx=10)


    def on_resource_change(self, index, region):
        cb = self.eu_labels[index] if region == 'eu' else self.non_eu_labels[index]
        default_role = self.default_roles_eu[index] if region == 'eu' else self.default_roles_noneu[index]

        selected_role = cb.get()

        if selected_role == default_role:
            cb.configure(background=self.auto_fill_color)  # Green
        else:
            cb.configure(background=self.user_edit_color)  # Blue

        self.update_rate_card(index, region)

        # After dropdown selection, compare the updated rate value too
        rate_entry = self.eu_rate_labels[index] if region == 'eu' else self.non_eu_rate_labels[index]
        default_rate = self.default_rate_eu[index] if region == 'eu' else self.default_rate_noneu[index]

        if rate_entry.get().strip() == default_rate:
            rate_entry.config(bg=self.auto_fill_color)
        else:
            rate_entry.config(bg=self.user_edit_color)


    def update_region_view(self):
        mode = self.region_mode.get()

        for i in range(23):

            # Determine default role
            if i < 8:
                default_role = "Junior"
            elif i < 19:
                default_role = "Specialist"
            else:
                default_role = "Expert"

            # ----- EU Part -----
            if mode in ["eu", "both"]:
                self.eu_dropdowns[i].config(state='normal')
                self.eu_dropdowns[i].delete(0, tk.END)
                self.eu_dropdowns[i].insert(0, "100" if mode == "eu" else "50")
                self.eu_dropdowns[i].config(bg=self.auto_fill_color)

                self.eu_labels[i].config(state='readonly')
                self.eu_labels[i].set(default_role)
                self.eu_labels[i].configure(background=self.auto_fill_color)

                self.update_rate_card(i, 'eu')
                self.eu_rate_labels[i].config(state='normal', disabledbackground="#e0e0e0", bg=self.auto_fill_color)

            else:  # non-eu mode
                self.eu_dropdowns[i].delete(0, tk.END)
                self.eu_dropdowns[i].insert(0, "0")
                self.eu_dropdowns[i].config(state='disabled', disabledbackground="#e0e0e0")
                self.eu_dropdowns[i].config(bg=self.auto_fill_color)
                self.eu_labels[i].config(state='disabled')
                self.eu_rate_labels[i].config(state='disabled')
                self.eu_rate_labels[i].config(disabledbackground="#e0e0e0")

            # ----- Non-EU Part -----
            if mode in ["non-eu", "both"]:
                self.non_eu_dropdowns[i].config(state='normal')
                self.non_eu_dropdowns[i].delete(0, tk.END)
                self.non_eu_dropdowns[i].insert(0, "100" if mode == "non-eu" else "50")
                self.non_eu_dropdowns[i].config(bg=self.auto_fill_color)
                

                self.non_eu_labels[i].config(state='readonly')
                self.non_eu_labels[i].set(default_role)
                self.non_eu_labels[i].configure(background=self.auto_fill_color)
                self.update_rate_card(i, 'non_eu')
                self.non_eu_rate_labels[i].config(state='normal', disabledbackground="#e0e0e0", bg=self.auto_fill_color)

            else:  # eu mode
                self.non_eu_dropdowns[i].delete(0, tk.END)
                self.non_eu_dropdowns[i].insert(0, "0")
                self.non_eu_dropdowns[i].config(state='disabled', disabledbackground="#e0e0e0")
                self.non_eu_dropdowns[i].config(bg=self.auto_fill_color)

                self.non_eu_labels[i].config(state='disabled')
                self.non_eu_rate_labels[i].config(state='disabled')
                self.non_eu_rate_labels[i].config(disabledbackground="#e0e0e0")


    # adding colors defaultly
    def on_user_edit(self, event):
        widget = event.widget
        current_value = widget.get().strip()

        try:
            index = None

            # EU Rate Card
            if widget in self.eu_rate_labels:
                index = self.eu_rate_labels.index(widget)
                default_val = self.default_rate_eu[index]

            # Non-EU Rate Card
            elif widget in self.non_eu_rate_labels:
                index = self.non_eu_rate_labels.index(widget)
                default_val = self.default_rate_noneu[index]

            # Efforts Field
            elif widget in self.complexity_entries:
                index = self.complexity_entries.index(widget)
                default_val = self.default_efforts[index]

            # EU %
            elif widget in self.eu_dropdowns:
                index = self.eu_dropdowns.index(widget)
                default_val = "100" if self.region_mode.get() == "eu" else "50" if self.region_mode.get() == "both" else "0"

            # Non-EU %
            elif widget in self.non_eu_dropdowns:
                index = self.non_eu_dropdowns.index(widget)
                default_val = "100" if self.region_mode.get() == "non-eu" else "50" if self.region_mode.get() == "both" else "0"

            # If index is matched and valid
            if index is not None and current_value == default_val:
                widget.config(bg=self.auto_fill_color)  # Green (default)
            else:
                widget.config(bg=self.user_edit_color)  # Blue (user-edited)

        except Exception as e:
            widget.config(bg=self.user_edit_color)


    def show_summary(self):
        try:
            self.sap.user_name = self.name_entry.get()
            self.sap.mode = self.mode.get()

            # Collect inputs from GUI into SAP object
            for i in range(23):
                self.sap.complexity_matrix[i][self.sap.mode] = float(self.complexity_entries[i].get())

                if i < 19 and self.inputs[i]:
                    self.sap.volume[i] = int(self.inputs[i].get())

                self.sap.eu[i] = int(self.eu_dropdowns[i].get())
                self.sap.non_eu[i] = int(self.non_eu_dropdowns[i].get())

                self.sap.eu_resource[i] = self.eu_labels[i].get()
                self.sap.non_eu_resource[i] = self.non_eu_labels[i].get()

                #  NEW LINES: Read values from editable Rate Card entries
                self.sap.eu_rate_card[i] = float(self.eu_rate_labels[i].get())
                self.sap.non_eu_rate_card[i] = float(self.non_eu_rate_labels[i].get())

            # Recalculate
            self.sap.calculate_efforts()
            self.sap.final_calculations()

            # Now read the summary values
            msg = (
                        f" Summary for: {self.sap.user_name}\n\n"
                        f"Total FTEs:     {self.sap.fte_summary:.2f}\n"
                        f"Total Cost:     ₹{self.sap.cost_summary:,.2f}\n"
                        f"Total Price:    ₹{self.sap.price_summary:,.2f}"
                  )   

            messagebox.showinfo("Estimation Summary", msg)

        except Exception as e:
            messagebox.showerror("Error", f"Failed to calculate summary:\n{e}")


    def fill_default_complexity(self):
        self.default_efforts.clear()
        for i in range(23):
            self.complexity_entries[i].delete(0, tk.END)
            default_value = str(self.sap.complexity_matrix[i][self.mode.get()])
            self.complexity_entries[i].insert(0, default_value)
            self.complexity_entries[i].config(bg=self.auto_fill_color)
            self.default_efforts.append(default_value)


    def update_rate_card(self, index, region):
        # Default rate card for rows 0–18
        default_rates = {
            "Junior": 22.71 if region == 'eu' else 11.21,
            "Senior": 37.8 if region == 'eu' else 17.82,
            "Specialist": 25.10 if region == 'eu' else 14.66,
            "Expert": 28.47 if region == 'eu' else 15.79
        }

        cb = self.eu_labels[index] if region == 'eu' else self.non_eu_labels[index]
        cb.configure(style="Blue.TCombobox")
        rate_entry = self.eu_rate_labels[index] if region == 'eu' else self.non_eu_rate_labels[index]
        role = cb.get()

        # Determine the correct rate card
        if index >= 19:
            row_offset = index - 19
            if region == 'eu':
                rate = self.sap.fte_rate_cards_eu[row_offset][role]
            else:
                rate = self.sap.fte_rate_cards_noneu[row_offset][role]
        else:
            rate = default_rates[role]

        # Update SAP role value
        if region == 'eu':
            self.sap.eu_resource[index] = role
        else:
            self.sap.non_eu_resource[index] = role

        # Fill rate entry box
        rate_entry.delete(0, tk.END)
        rate_entry.insert(0, str(rate))


    def focus_next(self, widget_list, index):
        if index + 1 < len(widget_list):
            next_widget = widget_list[index + 1]
            if next_widget:
                next_widget.focus_set()

    def export_excel(self):
        try:
            self.sap.user_name = self.name_entry.get()
            self.sap.mode = self.mode.get()

            for i in range(23):
                self.sap.complexity_matrix[i][self.sap.mode] = float(self.complexity_entries[i].get())
                if i < 19 and self.inputs[i]:
                    self.sap.volume[i] = int(self.inputs[i].get())

                self.sap.eu[i] = int(self.eu_dropdowns[i].get())
                self.sap.non_eu[i] = int(self.non_eu_dropdowns[i].get())

            self.sap.calculate_efforts()
            self.sap.final_calculations()
            
            # Exporting
            import subprocess
            from tkinter import messagebox

            excel_path = self.sap.fill_excel(region_mode=self.region_mode.get())
            folder_path = os.path.dirname(excel_path)

            if messagebox.askyesno("Export Successful", "Excel file exported successfully!\n\nDo you want to open the file location?"):
                if os.path.exists(folder_path):
                    subprocess.Popen(f'explorer "{folder_path}"')  
                else:
                    messagebox.showerror("Error", "Sorry, the folder could not be found.")

        except Exception as e:
            messagebox.showerror("Error", f"Failed to export Excel:\n{e}")


if __name__ == '__main__':
    try:
        root = tk.Tk()
        app = SAP_GUI(root)
        root.mainloop()
    except Exception as e:
        import traceback
        from tkinter import messagebox
        traceback.print_exc()  # Optional: keep it for console debugging
        messagebox.showerror("Unexpected Error", f"A critical error occurred:\n{str(e)}")

